import openpyxl as xl  # am adaugat si un aias "as xl" la capat pentru a apela mai usor functiile

workbook_object = xl.load_workbook('transactions.xlsx')
sheet = workbook_object['Sheet1']
cell = sheet['a1']  # se mai poate apela si astfel: cell = sheet.cell(1, 1)
# print(cell.value)
# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    # print(row)
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.max_row + 1
    corrected_price_cell.value = corrected_price