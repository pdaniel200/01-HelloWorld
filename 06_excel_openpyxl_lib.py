import openpyxl as xl  # am adaugat si un aias "as xl" la capat pentru a apela mai usor functiile
from openpyxl.chart import BarChart, Reference  # importa clasele BarChar si Reference din openpyxl pentru graficul din excel

# wb = xl.load_workbook(('transactions.xlsx'))
workbook_object = xl.load_workbook('transactions.xlsx')
sheet = workbook_object['Sheet1']
cell = sheet['a1']  # se mai poate apela si astfel: cell = sheet.cell(1, 1)
# print(cell.value)  # se afiseaza valoarea unei variabile care contine numele unei celule cu denumirea cell
# print(sheet.max_row)
# print(xl.__version__) # afisare versiune librarie
for row in range(2, sheet.max_row + 1):
    # print(row)
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = int(cell.value) * 0.9  # definim cum se calculeaza pretul nou
    corrected_price_cell = sheet.cell(row, 4)  # definim pozitia celulei cu pret nou
    # corrected_price_cell = sheet.max_row + 1
    corrected_price_cell.value = corrected_price  # setam valoare celulei cu pret nou, adica pozitia celulei cu pret nou are valoarea variabilei care contine pretul nou

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data((values))
sheet.add_chart(chart, "e2")

workbook_object.save('transactions2.xlsx')
